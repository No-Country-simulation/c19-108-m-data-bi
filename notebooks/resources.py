import zipfile
import numpy as np
import pandas as pd
import os
import missingno as msno
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import seaborn as sns
from collections import Counter
import re
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq

import warnings
warnings.filterwarnings(action= 'ignore')




def Imagen(path_relativo: str, size: tuple[int,int], path_absoluto:  None|str= None):

    """
    Esta funcion recibe un path relativo referente a una imagen (o un path absoluto), y el tamaño/dimensiones que
    debe tener en forma de tupla (width, height)

    devuelve la imagen pasada con las dimensiones obtenidas

    parameters: path_relativo (str), size (tuple[int,int]), path_absoluto(None,str)

    returns: None
    """

    import matplotlib.pyplot as plt
    import matplotlib.image as mpimg



    # Load image
    image = mpimg.imread(path_relativo)
    # Let the axes disappear
    plt.figure(figsize= size)
    plt.axis('off')
    # Plot image in the output
    image_plot = plt.imshow(image)



# Abre e instancia los ficheros (comprimidos en formato .zip o que esten en .csv)
def instanciar_data(path: str) -> str|pd.DataFrame:

    """
    Instancia un DataFrame a partir de un archivo CSV, ya sea directamente o desde un archivo ZIP.
    
    Args:
        path (str): La ruta del archivo CSV o ZIP que contiene el CSV.
        
    Returns:
        pd.DataFrame: El DataFrame instanciado a partir del archivo CSV.
        str: Mensaje de error si la ruta o el archivo no existen.
    """

    # Verificar si la ruta o el archivo existe
    if os.path.exists(path):
        # Evaluar si es un archivo .zip
        if path.lower().endswith('.zip'):
            # Instanciar el archivo comprimido en un objeto ZipFile
            with zipfile.ZipFile(path, 'r') as f:
                # Listar los archivos contenidos en el .zip
                file = f.namelist()
                if not file:
                    return 'El archivo .zip está vacío'
                
                # Abrir el archivo .csv que contiene
                with f.open(file[0], 'r') as p:
                    # Instanciar el .csv en un objeto DataFrame
                    df = pd.read_csv(p)
        else:
            # Leer el archivo .csv directamente
            with open(path, 'r', encoding='utf-8') as p:
                df = pd.read_csv(p)
        return df
    else:
        return 'El archivo o la ruta del archivo no existe'


#- INFORME PRELIMINAR

#- VERIFICAR EL TIPO DE DATOS.

def verificar_tipo_datos(df):

    '''
    Realiza un análisis de los tipos de datos y la presencia de valores nulos en un DataFrame.

    Esta función toma un DataFrame como entrada y devuelve un resumen que incluye información sobre
    los tipos de datos en cada columna, el porcentaje de valores no nulos y nulos, así como la
    cantidad de valores nulos por columna.

    Parameters:
        df (pandas.DataFrame): El DataFrame que se va a analizar.

    Returns:
        pandas.DataFrame: Un DataFrame que contiene el resumen de cada columna, incluyendo:
        - 'nombre_campo': Nombre de cada columna.
        - 'tipo_datos': Tipos de datos únicos presentes en cada columna.
        - 'no_nulos_%': Porcentaje de valores no nulos en cada columna.
        - 'nulos_%': Porcentaje de valores nulos en cada columna.
        - 'nulos': Cantidad de valores nulos en cada columna.
    '''

    mi_dict = {"nombre_campo": [], "tipo_datos": [], "no_nulos_%": [], "nulos_%": [], "nulos": []}

    for columna in df.columns:
        porcentaje_no_nulos = (df[columna].count() / len(df)) * 100
        mi_dict["nombre_campo"].append(columna)
        mi_dict["tipo_datos"].append(df[columna].apply(type).unique())
        mi_dict["no_nulos_%"].append(round(porcentaje_no_nulos, 2))
        mi_dict["nulos_%"].append(round(100-porcentaje_no_nulos, 2))
        mi_dict["nulos"].append(df[columna].isnull().sum())

    df_info = pd.DataFrame(mi_dict)

    return df_info



#- INFORME DEL DATA FRAME.

def informe_dataframe(dataframe: pd.DataFrame) -> None:

    """
    esta funcion obtiene un dataframe, y realiza un informe analizando y explorando algunas caracteristicas del
    dataframe centrandose principalmente en caracteristicas a nivel general de nuestro dataframe y realizando un procesamiento de
    algunos datos obteniendo metricas e informacion

    devuelve un informe que contiene:

    -Dimensiones del DataFrame
    -Numero de datos
    -Filas y Columnas
    -Tipo de columnas
    -Cantidad de registros duplicados
    -Metricas Generales

    Parameters: data (pandas.DataFrame).

    Returns: None.

    """

    df = dataframe

    print('INFORME PRELIMINAR SOBRE CARACTERISTICAS DEL DATASET:\n')
    print(f'--Dimensiones del DataFrame--\nFilas: {df.shape[0]}\nColumnas: {df.shape[1]}\n')
    print(f'--Numero de datos--\n{df[df.isna() == False].count().sum()}\n')
    print(f'--Filas y Columnas--\nFilas: muestra de indices-------> {list(df.index)[0:5]}  -----> Desde {list(df.index)[0]}  Hasta {list(df.index)[-1]}\nColumnas: {list(df.columns)}\n')
    print(f'--Tipo de columnas--\n{df.dtypes}')
    columnas= df.columns

    if 'hours' in columnas and 'attributes' not in columnas:
        print(f'--Cantidad de registros duplicados--\n{df.drop(columns=["hours"]).duplicated().sum()}\n')
    elif 'hours' not in columnas and 'attributes' in columnas:
        print(f'--Cantidad de registros duplicados--\n{df.drop(columns=["attributes"]).duplicated().sum()}\n')
    elif 'hours' in columnas and 'attributes' in columnas:
        print(f'--Cantidad de registros duplicados--\n{df.drop(columns=["hours","attributes"]).duplicated().sum()}\n')
    else:
        print(f'--Cantidad de registros duplicados--\n{df.duplicated().sum()}\n')


    # print(f'--Estadisticos preliminares generales--\n{df.describe()}\n')

    return ('~'*50)+'oo'+('~'*50)



#----------------------------------------------------------------------------------------------------

#- INFORME DE COLUMNA.

def informe_columna(df: None|pd.DataFrame, columna: None|str) -> None:

    """
    esta funcion obtiene un dataframe y el nombre de una de sus columnas, y realiza un informe analizando y explorando algunas caracteristicas de
    la feature, centrandose principalmente en caracteristicas a nivel general y realizando un procesamiento de 
    algunos datos obteniendo metricas e informacion

    Dependiendo el tipo de dato contenido en la feature/columna, devolvera informacion ligeramente diferente:

    Para tipo object:

    -Numero de datos nulos
    -Cantidad de valores unicos en la columna
    -Valores unicos en la columna (Primeros 5 valores, en caso de exceder los 5, en caso contrario, devuelve todos los valores unicos)
    -Moda de la columna
    -Distribucion de frecuencias

    Para tipo datetime64[ns]:

    -Numero de datos nulos
    -Cantidad de valores unicos en la columna
    -Valores unicos en la columna (una muestra de 4 valores como ejemplo, y el rango que abarcan (desde que valor hasta que valor))
    -Moda de la columna
    -Distribucion de frecuencias
    -Valor maximo y minimo

    Para tipo numerico (int, float):

    -Numero de datos nulos
    -Valores unicos en la columna (una muestra de 5 valores como ejemplo, y el rango que abarcan (desde que valor hasta que valor))
    -Moda de la columna
    -Estadisticos Principales de la columna
    -Valores extremos
    -Distribucion de frecuencias
    -Valor maximo y minimo


    Parameters: data (pandas.DataFrame), columna (str).

    Returns: None.
    
    """

    data = df[columna]
    
    # print(f'Informe preliminar sobre la columna/feature {columna}:\n')
    print(f'INFORME PRELIMINAR SOBRE LA COLUMNA/FEATURE {columna}:\n')
    if data.dtype == 'object' or data.dtype == 'bool':
        print(f'--Numero de datos nulos--\n{data.isna().sum()}\n')
        print(f'--Cantidad de valores unicos en la columna--\n{data.describe()[1]}\n')

        if len(data.unique()) > 5:
            print(f'--Valores unicos en la columna (Primeros 5 valores)--\n{data.unique()[0:5]}\n')
        else:
            print(f'--Valores unicos en la columna--\n{data.unique()}\n')
            
        print(f'--Moda de la columna especificada--\nValor modal -----> {data.describe()[2]}\nFrecuencia acumulada ------> {data.describe()[3]}\n')
        print(f'--Distribucion de frecuencias (primeros valores con mayor cantidad de frecuencias)--\n {data.value_counts().nlargest(3)}\n')
        print('-'*120)
        print('-'*120)
    elif data.dtype == 'datetime64[ns]':
        print(f'--Numero de datos nulos--\n{data.isna().sum()}\n')
        print(f'--Cantidad de valores unicos en la columna--\n{data.describe()[1]}\n')
        ## En el print siguinte, realizamos un formateo de los valores de la columna, ya que la salida predeterminada (el output) agrega otros valores que hacen la intrepretacion mas dificil e incomoda
        print(f'--Valores unicos en la columna--\nEj: {data.dt.strftime("%Y-%m-%d").unique()[0:3]}  -----> Desde {list(data.dt.strftime("%Y-%m-%d").unique())[0]}  Hasta {list(data.dt.strftime("%Y-%m-%d").unique())[-1]}\n')
        print(f'--Moda de la columna especificada--\nValor modal -----> {data.describe()[2]}\nFrecuencia acumulada ------> {data.describe()[3]}\n')
        print(f'--Distribucion de frecuencias (primeros valores con mayor cantidad de frecuencias)--\n {data.value_counts().nlargest(3)}\n')
        print(f'--Valor maximo y minimo--\nMaximo: {data.max()}\nMinimo: {data.min()}\n')
        print('-'*120)
        print('-'*120)
    else:
        print(f'--Numero de datos nulos--\n{data.isna().sum()}\n')
        print(f'--Valores unicos en la columna--\nEj: {data.unique()[0:5]}  -----> Desde {list(data.unique())[0]}  Hasta {list(data.unique())[-1]}\n')
        print(f'--Estadisticos Principales de la columna--\nMedia: {round(data.mean(),2)}\nDesviacion Estandar: {round(data.std(),2)}\nPrimer cuartil: {data.quantile(0.25)}\nMediana: {data.median()}\nTercer cuartil: {data.quantile(0.75)}\n')
        print(f'--Valores extremos--\nValor maximo: {data.max()}\nValor minimo: {data.min()}\n')
        print(f'--Distribucion de frecuencias (primeros valores con mayor cantidad de frecuencias)--\n {data.value_counts().nlargest(3)}\n')
        print(f'--Valor maximo y minimo--\nMaximo: {data.max()}\nMinimo: {data.min()}\n')
        print('-'*120)
        print('-'*120)
    return


#----------------------------------------------------------------------------------------------------



#- VISUALIZACIÓN DE NULOS.

# Subplot

def visualizar_nulos(df: pd.DataFrame|None) -> None:

    """
    Toma como parametro un dataframe pandas y realiza la construccion de 4 graficos principales en
    base al analisis de sus datos con valor np.nan o nulos:

    Matrix Plot: Una cuadrícula donde cada fila representa una observación y cada columna representa una variable en el DataFrame.
    Los valores nulos están marcados en blanco en la cuadrícula.

    Bar Plot: Un gráfico de barras vertical que muestra la cantidad de valores nulos en cada columna del DataFrame.
    Cada barra vertical representa una columna, y la altura de la barra indica la cantidad de valores nulos en esa columna.

    Heatmap: Similar a la matriz, pero con colores para resaltar las áreas con más valores nulos.
    Los valores nulos pueden estar representados con colores más oscuros para destacar las áreas con más ausencias de datos.

    Dendrogram: Un diagrama de árbol que agrupa las variables y observaciones basadas en la similitud de los patrones de valores nulos.
    Las ramas del árbol se fusionan según la similitud de los patrones de valores nulos entre las variables y observaciones.

    Parameters: df (pandas.DataFrame)

    returns: None

    """
    

    fig, axes = plt.subplots(1, 2, figsize=(15, 10))

    msno.matrix(df, ax=axes[0])
    msno.bar(df, ax=axes[1])  


    axes[0].set_title('Matrix Plot')
    axes[1].set_title('Bar Plot')


    plt.tight_layout()
    plt.show()


    fig, axes = plt.subplots(1, 2, figsize=(15, 10))
    msno.heatmap(df, ax=axes[0])
    msno.dendrogram(df, ax=axes[1])


    axes[0].set_title('Heatmap')
    axes[1].set_title('Dendrogram')

    plt.tight_layout()
    plt.show()




#----------------------------------------------------------------------------------------------------



#- VALORES ATÍPICOS, EXTREMOS Y OUTLIERS.



# Columnas Cualitativas.

import pandas as pd
import matplotlib.pyplot as plt
from collections import Counter
import re

def analisis_frecuencia_palabras(dataframe: pd.DataFrame):
    """
    Toma como parámetro un dataframe pandas, selecciona las columnas tipo 'object', 
    realiza una serie de normalización y procesamiento de los datos, y genera un gráfico 
    de barras por cada columna seleccionada que representa las palabras más frecuentes en la columna.

    parameters: dataframe (pandas.DataFrame)

    returns: None
    """
    if dataframe is None:
        raise ValueError("El DataFrame proporcionado es None.")
        
    # Seleccionar columnas de tipo 'object'
    columnas_cualitativas = dataframe.select_dtypes(include=['object']).columns.tolist()

    for columna in columnas_cualitativas:
        # Asegurarse de que todos los valores son de tipo str y manejar valores nulos
        dataframe[columna] = dataframe[columna].astype(str).fillna('')

    for columna in columnas_cualitativas:
        # Procesamiento de texto
        texto_columna = dataframe[columna].str.lower().str.replace(r'[^a-zA-Z\s]', '', regex=True).str.split()
        palabras_columna = [word for sublist in texto_columna for word in sublist]
        frecuencia_palabras = Counter(palabras_columna)
        palabras_mas_frecuentes = frecuencia_palabras.most_common(10)
        
        # Separar palabras y sus frecuencias
        if palabras_mas_frecuentes:
            palabras, frecuencias = zip(*palabras_mas_frecuentes)
        else:
            palabras, frecuencias = [], []
        
        # Generar gráfico de barras
        plt.figure(figsize=(10, 6))
        plt.bar(palabras, frecuencias)
        plt.xticks(rotation=90)
        plt.xlabel('Palabra')
        plt.ylabel('Frecuencia')
        plt.title(f'Palabras más frecuentes en la columna {columna}')
        plt.show()

# Ejemplo de uso
# df = pd.read_csv('tu_archivo.csv')
# analisis_frecuencia_palabras(df)


#----------------------------------------------------------------------------------------------------


def boxplots_numericas(dataframe: pd.DataFrame|None) -> None:

    """
    Toma como parametro un pandas.DataFrame y devuelve un boxplot de
    cada columna de tipo numerico (varibles cuantitativas)

    parameters: dataframe (pd.DataFrame|None)

    returns: None
    """
    if dataframe is None:
        raise ValueError("El DataFrame proporcionado es None.")
    df_numericas = dataframe.select_dtypes(include=['number'])
    for columna in df_numericas.columns:
        sns.boxplot(x=dataframe[columna])
        plt.title(f'Diagrama de caja para {columna}')
        plt.show()

#----------------------------------------------------------------------------------------------------


# - VALORES DUPLICADOS.

def hist_duplicados(dataframe: str|pd.DataFrame) -> None:

    """
    Esta funcion toma como parametro un pandas.DataFrame.
    
    Da salida a un grafico de barras/histograma que refleja la cantidad de registros duplicados y no duplicados
    presentes en el dataframe

    En el Caso que no se registren duplicados, dara salida a un mensaje confirmandolo

    parameters: dataframe(str|pandas.DataFrame)

    returns: None
    """

    columnas= dataframe.columns

    if 'hours' in columnas and 'attributes' not in columnas:
        df_duplicates= dataframe.drop(columns=["hours"]).duplicated().value_counts()
    elif 'hours' not in columnas and 'attributes' in columnas:
        df_duplicates= dataframe.drop(columns=["attributes"]).duplicated().value_counts()
    elif 'hours' in columnas and 'attributes' in columnas:
        df_duplicates= dataframe.drop(columns=["hours","attributes"]).duplicated().value_counts()
    else:
        df_duplicates= dataframe.duplicated().value_counts()
        
    df_duplicates = pd.DataFrame({'Duplicados': df_duplicates.index, 'Frecuencia': df_duplicates.values})
    if not df_duplicates.empty: 
        sns.barplot(x='Duplicados', y='Frecuencia', data=df_duplicates)
        plt.title('Conteo de Frecuencias de Registros Duplicados')
        plt.xlabel('Duplicados')
        plt.ylabel('Frecuencia')
        plt.xticks([0, 1], ['No Duplicados', 'Duplicados'])
        plt.show()
    else:
        print("No se encontraron registros duplicados.")



#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------




#- PARA ANÁLISIS DESCRIPTIVO

#----------------------------------------------------------------------------------------------------

def descripcion_distribucion(df: pd.DataFrame):
    '''
    Crea una serie de gráficos de barras y de torta que muestran la distribución de los valores de la feature/columna.

    También imprime una serie de valores y métricas básicas en referencia a los valores y distribución de valores de la columna/feature

    Parameters:
        df (pandas.DataFrame): El DataFrame que contiene los datos de la recopilación de features de ICFES.

    Returns:
        None
    '''

    columnas = df.columns

    for columna in columnas:
        muestra = df[columna].value_counts().nlargest(10)

        plt.figure(figsize=(20, 10))

        if df[columna].dtype == 'object' or df[columna].dtype == 'datetime64[ns]':
            # Gráfico de barras
            plt.subplot(1, 2, 1)
            ax = sns.barplot(x=muestra.index, y=muestra.values, data=muestra.to_frame().reset_index())
            ax.set_title(f'Gráfico de distribución de frecuencias (primeros 10 valores) {columna}')
            ax.set_xlabel(columna)
            ax.set_ylabel('Cantidad de frecuencias')
            plt.xticks(rotation=45)

            # Gráfico de torta para las proporciones representativas de los datos
            plt.subplot(1, 2, 2)
            try:
                plt.pie(x=muestra, labels=muestra.index, shadow=True, autopct='%1.1f%%')
            except ValueError:
                muestra = muestra[:5]  # Tomar solo las 5 categorías principales
                plt.pie(x=muestra, labels=muestra.index, shadow=True, autopct='%1.1f%%')
            plt.title('Distribución Relativa de Valores (primeros 5)')
            plt.xlabel(columna)
            plt.xticks(rotation=45)
            plt.grid()
        else:
            # Diagrama de caja para columnas numéricas
            plt.subplot(1, 2, 1)
            sns.boxplot(x=df[columna])
            plt.title(f'Diagrama de caja para {columna}')


            # Grafico de densidad para representar la distribucion de frecuencias
            plt.subplot(1, 2, 2)
            sns.kdeplot(df[columna], shade= True)
            plt.title(f'Gráfico de Densidad para {columna}')
            plt.xlabel(columna)
            plt.ylabel('frecuencias')
            plt.xticks(rotation=45)
            plt.grid()

        # Ajustamos los graficos 
        plt.tight_layout()
        # Mostrar el gráfico
        plt.show()

        # Mostrar algunas estadísticas básicas
        print('\n')
        print('Estadísticas Básicas')
        print('\n')
        print(df[columna].describe())
        print('-' * 160)
        print('-' * 160)
        print('\n')



#----------------------------------------------------------------------------------------------------

def conversion_de_tablas_dinamicas(path: str, n_hoja: int, start_row: int, start_col: str, etiqueta_año: int):


    # Cargar el archivo Excel
    file_path = path  # path del fichero.xlsx
    workbook = openpyxl.load_workbook(file_path, data_only=True) # instancio la carga del fichero en una variable

    # Iterar sobre cada hoja del archivo
    for idx, sheet_name in enumerate(workbook.sheetnames, start=1): # itero sobre cada hoja tomando como referencia las etiquetas de cada una con el metodo sheetnames
        
        if idx != n_hoja:  # Solo procesar la tercera hoja
            continue

        sheet = workbook[sheet_name] # instancio la hoja
        
        # Identificar la tabla dinámica (puedes ajustar este criterio según tu necesidad)
        # Aquí asumimos que la tabla dinámica empieza en la fila 12 y columna A (cambiar según sea necesario)
        pivot_start_row = start_row
        pivot_start_col = start_col

        
        # Extraer la tabla dinámica
        pivot_data = []
        for row in sheet.iter_rows(min_row=pivot_start_row, min_col=openpyxl.utils.cell.column_index_from_string(pivot_start_col), values_only=True):
            pivot_data.append(row)
        
        # Convertir a DataFrame de pandas
        pivot_df = pd.DataFrame(pivot_data[1:], columns = pivot_data[0])

        archivo= f'../data/raw/{sheet_name}_{etiqueta_año}.parquet'
        pq.write_table(pa.Table.from_pandas(pivot_df), archivo)
        

    print(f'Nombre de dataframe: {sheet_name} {etiqueta_año}')
    return pivot_df

#----------------------------------------------------------------------------------------------------

def extraccion_de_tablas(path: str, n_hoja: int, start_row: int, start_col: str, end_row: int, end_col: str, etiqueta_tema: str):

    # Cargar el archivo Excel
    file_path = path  # path del fichero.xlsx
    workbook = openpyxl.load_workbook(file_path, data_only=True) # instancio la carga del fichero en una variable

    # Iterar sobre cada hoja del archivo
    for idx, sheet_name in enumerate(workbook.sheetnames, start=1): # itero sobre cada hoja tomando como referencia las etiquetas de cada una con el metodo sheetnames
        
        if idx != n_hoja:  # Solo procesar la tercera hoja
            continue

        sheet = workbook[sheet_name] # instancio la hoja
        
        # Identificar la tabla dinámica (puedes ajustar este criterio según tu necesidad)
        # Aquí asumimos que la tabla dinámica empieza en la fila 12 y columna A (cambiar según sea necesario)
        pivot_start_row = start_row
        pivot_end_row = end_row
        pivot_start_col = start_col
        pivot_end_col = end_col
        
        # Extraer la tabla dinámica
        pivot_data = []
        for row in sheet.iter_rows(min_row=pivot_start_row, max_row= pivot_end_row, min_col=openpyxl.utils.cell.column_index_from_string(pivot_start_col), max_col= openpyxl.utils.cell.column_index_from_string(pivot_end_col), values_only=True):
            pivot_data.append(row)
        
        #transformamos la lista de listas en un array para poder manejar el index
        pivot_data = np.array(pivot_data)

        # Convertir a DataFrame de pandas
        pivot_df = pd.DataFrame(pivot_data[1:, 1:], columns=pivot_data[0, 1:], index=pivot_data[1:, 0])

        # Reemplazar valores no numéricos con NaN
        pivot_df.replace({'n.a.': np.nan, '': np.nan}, inplace=True)
    
        # Convertir columnas a numéricas si es posible
        pivot_df = pivot_df.apply(pd.to_numeric, errors='coerce')

        archivo= f'../data/raw/{sheet_name}_{etiqueta_tema}.parquet'
        pq.write_table(pa.Table.from_pandas(pivot_df, preserve_index= True), archivo)
        

    print(f'Nombre de dataframe: {sheet_name} {etiqueta_tema}')
    return pivot_df

#----------------------------------------------------------------------------------------------------


# Renombrar columnas duplicadas
def make_unique(column_names):
    seen = set()
    for item in column_names:
        counter = 1
        new_item = item
        while new_item in seen:
            new_item = f"{item}_{counter}"
            counter += 1
        seen.add(new_item)
        yield new_item

#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------

def main():
    
    return

if __name__ == '__main__':
    main()